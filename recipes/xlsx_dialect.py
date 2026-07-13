#!/usr/bin/env python3
"""OSMS Excel dialect (range model, "Modell A").

One formula = one KPI value over a sheet named `data` (header in row 1, columns in
catalog order). Parameters live on a sheet `result` (B1=scope, B2=period start,
B3=period end). Fail-closed via NA(); percentiles via explicit nearest rank
(SMALL + CEILING) so no interpolating estimator is involved. Functions are limited
to the Excel 2007 set (COUNTIFS/SUMIFS/SUMPRODUCT/SMALL/MEDIAN/AVERAGE/COUNT/
CEILING/MAX/MIN/IF/N) so the same formula runs in Excel, LibreOffice Calc and
Google Sheets.

This module is the single source of truth: gen_recipes renders the display snippet
from build_spec, and both gates.py (formulas engine) and excel_runner.py
(LibreOffice engine) rebuild the executable workbook from the same spec.
"""
import re, os, subprocess, datetime as _dt

DUR_PAT = re.compile(r"([a-z_]+_at)\s*[-\u2212]\s*([a-z_]+_at)")
RATIO_PAT = re.compile(r"^\(?\s*([A-Za-z_][A-Za-z0-9_]*)\s*/\s*([A-Za-z_][A-Za-z0-9_]*)\s*\)?\s*(\*\s*100)?\s*$")

STD001_W = [("STD-006",0.25),("STD-014",0.08),("STD-015",0.06),("STD-013",0.04),("STD-029",0.02),
            ("SOC-002",0.06),("SOC-003",0.06),("SOC-004",0.05),("SOC-009",0.03),("STD-008",0.15),
            ("STD-007",0.04),("STD-053",0.035),("STD-054",0.025),("STD-011",0.10)]

P1, P2, P3 = "result!$B$1", "result!$B$2", "result!$B$3"   # params, referenced from data-sheet helpers


def LETTER(i):
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26); s = chr(65 + r) + s
    return s

def _cmap(cols): return {f: LETTER(i + 1) for i, f in enumerate(cols)}
def _rng(col): return "data!%s$2:%s$100000" % (col, col)

def _ascii(s):
    """Fold card-derived hook text to plain ASCII and neutralise double quotes so it
    can be embedded safely inside an N("...") documentation term."""
    s = (str(s).replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'")
         .replace("\u2019", "'").replace("\u201c", "'").replace("\u201d", "'")
         .replace("\u00b7", "-").replace("\u2026", "...").replace('"', "'"))
    return "".join(ch if ord(ch) < 127 else "?" for ch in s)


# German (de-DE) Excel localisation: same formula, localised function names, ";" argument
# separator and "," decimal. Only names that differ from en-US are listed; MEDIAN/MAX/MIN/ABS/N
# are identical in de-DE. Applied longest-first, before "(" only, with string literals protected.
_FUNC_DE = [("COUNTIFS", "Z\u00c4HLENWENNS"), ("SUMPRODUCT", "SUMMENPRODUKT"), ("SUMIFS", "SUMMEWENNS"),
            ("AVERAGE", "MITTELWERT"), ("CEILING", "OBERGRENZE"), ("SMALL", "KKLEINSTE"),
            ("COUNT", "ANZAHL"), ("SUM", "SUMME"), ("IF", "WENN"), ("AND", "UND"), ("OR", "ODER"),
            ("NA", "NV")]
_FUNC_EN = [(de, en) for en, de in _FUNC_DE]   # inverse, for the round-trip check

# German documentation phrases for the de-DE rendering. These are the fixed scaffolding
# labels emitted inside N("...") documentation terms; card-derived hook fragments and all
# data-value literals ("critical", "validated", ...) stay English. Longest-first so nested
# phrases ("denominator population" before "population") resolve correctly.
_NDOC = [
    ("range-model skeleton, non-normative", "Range-Modell-Skelett, nicht normativ"),
    ("add your population/condition criteria pairs to the two COUNTIFS",
     "Kriterien-Paare f\u00fcr Population/Bedingung in die beiden COUNTIFS erg\u00e4nzen"),
    ("a count of 0 is a valid result; n/a only when the source itself is unavailable",
     "Ein Z\u00e4hlwert von 0 ist ein g\u00fcltiges Ergebnis; k.A. nur wenn die Quelle selbst nicht verf\u00fcgbar ist"),
    ("delta = current-period quantity minus previous-period quantity",
     "Delta = Gr\u00f6\u00dfe der aktuellen Periode minus Gr\u00f6\u00dfe der Vorperiode"),
    ("map the period split and the quantity: ", "Periodenaufteilung und Gr\u00f6\u00dfe mappen: "),
    ("duration expression: map per the card formula", "Dauer-Ausdruck: gem\u00e4\u00df Kartenformel mappen"),
    ("denominator population: ", "Nenner-Population: "),
    ("numerator condition: ", "Z\u00e4hler-Bedingung: "),
    ("population: ", "Population: "),
]


def _split_strings(f):
    return re.split(r'("(?:[^"]*)")', f)       # odd indices are quoted string literals


def to_de_de(formula):
    """Render an en-US formula as paste-ready de-DE (German Excel) text: localised function
    names, ';' separator, ',' decimal, and German N() documentation. Data-value literals and
    card-derived hook fragments stay English."""
    out = []
    for i, seg in enumerate(_split_strings(formula)):
        if i % 2 == 1:                          # string literal: translate doc phrases only
            for en, de in _NDOC:
                seg = seg.replace(en, de)
            out.append(seg); continue
        for en, de in _FUNC_DE:
            seg = re.sub(r"\b" + en + r"(?=\()", de, seg)
        seg = re.sub(r"\bTRUE\b", "WAHR", seg)
        seg = seg.replace(",", ";")            # argument separators
        seg = re.sub(r"(?<=\d)\.(?=\d)", ",", seg)   # decimal points
        out.append(seg)
    return "".join(out)


def _to_en_from_de(formula):
    """Inverse of to_de_de, used only to prove the de-DE rendering is faithful."""
    out = []
    for i, seg in enumerate(_split_strings(formula)):
        if i % 2 == 1:                          # string literal: docs back to English
            for en, de in _NDOC:
                seg = seg.replace(de, en)
            out.append(seg); continue
        seg = re.sub(r"(?<=\d),(?=\d)", ".", seg)    # decimals back first
        seg = seg.replace(";", ",")
        seg = re.sub(r"\bWAHR\b", "TRUE", seg)
        for de, en in _FUNC_EN:
            seg = re.sub(r"\b" + de + r"(?=\()", en, seg)
        out.append(seg)
    return "".join(out)


# ============ spec builders (each returns dict or None) ============
# spec = {columns, params:[..], helpers:[(name, formula_template_with_{r})], outputs:{key:formula}, note}

def _ratio_concrete_std016(cols):
    c = _cmap(cols); r = lambda f: _rng(c[f])
    den = ('COUNTIFS(%s,$B$1,%s,"critical",%s,TRUE,%s,">="&$B$2,%s,"<"&$B$3)'
           % (r("scope_id"), r("severity"), r("internet_facing"), r("due_at"), r("due_at")))
    num = ('SUMPRODUCT((%s=$B$1)*(%s="critical")*(%s=TRUE)*(%s>=$B$2)*(%s<$B$3)*(%s<=%s)*(%s="validated"))'
           % (r("scope_id"), r("severity"), r("internet_facing"), r("due_at"), r("due_at"),
              r("remediated_at"), r("due_at"), r("validation_status")))
    return dict(columns=cols, params=["scope", "ps", "pe"], helpers=[],
                outputs={"sla_compliance_pct": "=IF(%s=0,NA(),100*%s/%s)" % (den, num, den)},
                note="Population: in-scope critical internet-facing findings whose due date is in the period; "
                     "numerator: remediated on or before due and validated. Empty population -> n/a, never 0.")

def _ratio_skeleton(cid, cols, hooks):
    c = _cmap(cols)
    if "scope_id" not in c:
        return None
    sc = _rng(c["scope_id"])
    base = ("=IF(COUNTIFS(%s,$B$1)=0,NA(),"
            "100*COUNTIFS(%s,$B$1)/COUNTIFS(%s,$B$1))" % (sc, sc, sc))
    base += ('+N("%s : range-model skeleton, non-normative")' % cid)
    base += ('+N("denominator population: %s")' % _ascii(hooks["den"]))
    base += ('+N("numerator condition: %s")' % _ascii(hooks["num"]))
    base += '+N("add your population/condition criteria pairs to the two COUNTIFS")'
    return dict(columns=cols, params=["scope"], helpers=[],
                outputs={"value": base},
                note="Skeleton: the scope filter and fail-closed guard are generated; extend the two "
                     "COUNTIFS with the card's population (denominator) and condition (numerator) criteria.")

def _duration(cols, start, end, period_on_start, concrete):
    c = _cmap(cols)
    if "scope_id" not in c:
        return None
    if concrete and not ({start, end} <= set(c)):
        concrete = False
    hcol = LETTER(len(cols) + 1); H = _rng(hcol)
    S = "$%s{r}" % c["scope_id"]
    if concrete:
        A = "$%s{r}" % c[start]; B = "$%s{r}" % c[end]
        cond = 'AND(%s=%s,%s<>"",%s<>"",%s>=%s%s)' % (
            S, P1, A, B, B, A, (",%s>=%s,%s<%s" % (A, P2, A, P3) if period_on_start else ""))
        helper = ("d_h", '=IF(%s,(%s-%s)*24,"")' % (cond, B, A))
    else:
        helper = ("d_h", '=IF(%s=%s,0,"")+N("duration expression: map per the card formula")' % (S, P1))
    return dict(columns=cols, params=(["scope", "ps", "pe"] if period_on_start else ["scope"]),
                helpers=[helper],
                outputs={
                    "valid_cases": "=COUNT(%s)" % H,
                    "p50_h": "=IF(COUNT(%s)=0,NA(),MEDIAN(%s))" % (H, H),
                    "p90_h": "=IF(COUNT(%s)=0,NA(),SMALL(%s,CEILING(0.9*COUNT(%s),1)))" % (H, H, H),
                    "mean_h_supplementary": "=IF(COUNT(%s)=0,NA(),AVERAGE(%s))" % (H, H)},
                note=("Helper column d_h holds the per-case duration in hours; P50 is the true median, "
                      "P90 the exact nearest rank (CEILING(0.9*n)). Empty case base -> n/a, never 0."
                      if concrete else
                      "Skeleton: map the duration expression in helper column d_h; the percentile block is "
                      "generated (P50 median, P90 nearest rank)."))

def _count_skeleton(cid, cols, hook):
    c = _cmap(cols)
    if "scope_id" not in c: return None
    sc = _rng(c["scope_id"])
    f = ('=COUNTIFS(%s,$B$1)+N("population: %s")'
         '+N("a count of 0 is a valid result; n/a only when the source itself is unavailable")' % (sc, _ascii(hook)))
    return dict(columns=cols, params=["scope"], helpers=[], outputs={"value": f},
                note="Skeleton: add the card's population criteria to COUNTIFS. A count of 0 is a true 0.")

def _delta_skeleton(cid, cols, hook):
    c = _cmap(cols)
    if "scope_id" not in c: return None
    sc = _rng(c["scope_id"])
    f = ('=IF(COUNTIFS(%s,$B$1)=0,NA(),0)+N("map the period split and the quantity: %s")'
         '+N("delta = current-period quantity minus previous-period quantity")' % (sc, _ascii(hook)))
    return dict(columns=cols, params=["scope"], helpers=[], outputs={"value": f},
                note="Skeleton: replace 0 with the current-minus-previous period quantity; empty base -> n/a.")

def _composite(cols, k):
    c = _cmap(cols); r = lambda f: _rng(c[f])
    need = {"scope_id", "period_start", "period_end", "weight", "subscore_value"}
    if not (need <= set(c)): return None
    M = "(%s=$B$1)*(%s=$B$2)*(%s=$B$3)" % (r("scope_id"), r("period_start"), r("period_end"))
    n = "SUMPRODUCT(%s)" % M
    wsum = "SUMPRODUCT(%s*%s)" % (M, r("weight"))
    oob = "SUMPRODUCT(%s*((%s<0)+(%s>100)))" % (M, r("subscore_value"), r("subscore_value"))
    score = "SUMPRODUCT(%s*%s*%s)" % (M, r("weight"), r("subscore_value"))
    gate = "OR(%s<>%d,ABS(%s-1)>0.001,%s>0)" % (n, k, wsum, oob)
    return dict(columns=cols, params=["scope", "ps", "pe"], helpers=[],
                outputs={"value": "=IF(%s,NA(),%s)" % (gate, score)},
                note="Weighted sum of pre-scored 0-100 subscore rows; gated on component count, weight sum = 1.0 "
                     "and value bounds. Any violation -> n/a, never a fabricated score.")

def _std001(cols):
    c = _cmap(cols); r = lambda f: _rng(c[f])
    if not ({"child_card_id", "posture_score", "penalty_value", "scope_id"} <= set(c)): return None
    terms = " + ".join('%s*SUMIFS(%s,%s,"%s",%s,$B$1)'
                       % (w, r("posture_score"), r("child_card_id"), k, r("scope_id")) for k, w in STD001_W)
    nleaf = " + ".join('COUNTIFS(%s,$B$1,%s,"%s")' % (r("scope_id"), r("child_card_id"), k) for k, _ in STD001_W)
    pen = ('SUMIFS(%s,%s,"STD-001a",%s,$B$1)+SUMIFS(%s,%s,"STD-001b",%s,$B$1)'
           % (r("penalty_value"), r("child_card_id"), r("scope_id"),
              r("penalty_value"), r("child_card_id"), r("scope_id")))
    pn = ('COUNTIFS(%s,$B$1,%s,"STD-001a")+COUNTIFS(%s,$B$1,%s,"STD-001b")'
          % (r("scope_id"), r("child_card_id"), r("scope_id"), r("child_card_id")))
    oob = ('SUMPRODUCT((%s=$B$1)*(%s<>"STD-001a")*(%s<>"STD-001b")*((%s<0)+(%s>100)))'
           % (r("scope_id"), r("child_card_id"), r("child_card_id"), r("posture_score"), r("posture_score")))
    gate = "OR((%s)<>14,(%s)<>2,%s>0)" % (nleaf, pn, oob)
    return dict(columns=cols, params=["scope"], helpers=[],
                outputs={"value": "=IF(%s,NA(),MAX(0,MIN(100,(%s)-(%s))))" % (gate, terms, pen)},
                note="Two-level composite: 14 pre-multiplied leaf weights (sum to 1.0) times posture scores, "
                     "minus the two penalty rows, clamped to 0..100. Gates: 14 leaves present once, both penalties "
                     "present, scores 0-100 -> otherwise n/a.")

def _res007(cols):
    c = _cmap(cols); r = lambda f: _rng(c[f])
    need = {"scope_id", "p90_loss_exposure_eur", "policy_limit_eur", "deductible_eur", "exclusion_does_not_apply"}
    if not (need <= set(c)): return None
    hcol = LETTER(len(cols) + 1); H = _rng(hcol)
    cc = lambda f: "$%s{r}" % c[f]
    helper = ("uncovered_eur",
              '=IF(%s<>%s,"",MAX(%s-MAX(MIN(%s,%s)-%s,0)*%s,0))'
              % (cc("scope_id"), P1, cc("p90_loss_exposure_eur"),
                 cc("policy_limit_eur"), cc("p90_loss_exposure_eur"),
                 cc("deductible_eur"), cc("exclusion_does_not_apply")))
    n = "COUNTIFS(%s,$B$1)" % r("scope_id")
    bad = ('SUMPRODUCT((%s=$B$1)*((%s<0)+(%s<0)+(%s<0)+((%s<>0)*(%s<>1))+(%s="")+(%s="")+(%s="")+(%s="")))'
           % (r("scope_id"), r("p90_loss_exposure_eur"), r("policy_limit_eur"), r("deductible_eur"),
              r("exclusion_does_not_apply"), r("exclusion_does_not_apply"),
              r("p90_loss_exposure_eur"), r("policy_limit_eur"), r("deductible_eur"), r("exclusion_does_not_apply")))
    return dict(columns=cols, params=["scope"], helpers=[helper],
                outputs={"value": "=IF(OR(%s=0,%s>0),NA(),SUM(%s))" % (n, bad, H)},
                note="Helper column uncovered_eur = max(P90 exposure - creditable coverage, 0) per scenario "
                     "(limit erosion, the card's declared assumption); value = the sum over scenarios. Empty or "
                     "any invalid row -> n/a, never 0.")

def _log015(cols):
    c = _cmap(cols); r = lambda f: _rng(c[f])
    need = {"criticality_tier", "first_approval_date", "exception_scope", "compensating_control_ref",
            "status", "period_end", "scope_id"}
    if not (need <= set(c)): return None
    hcol = LETTER(len(cols) + 1); H = _rng(hcol)
    rc = lambda f: "$%s{r}" % c[f]
    wc = 'IF(%s="tier-0",5,IF(%s="tier-1",3,1))' % (rc("criticality_tier"), rc("criticality_tier"))
    age = "(%s-%s)" % (rc("period_end"), rc("first_approval_date"))
    wa = "IF(%s<=30,1,IF(%s<=90,2,3))" % (age, age)
    we = ('IF(%s="individual_asset",1,IF(%s="asset_group_or_service",2,3))'
          % (rc("exception_scope"), rc("exception_scope")))
    wk = 'IF(%s<>"",1,2)' % rc("compensating_control_ref")
    helper = ("weighted_exposure",
              '=IF(AND(%s="open",%s=%s),%s*%s*%s*%s,"")'
              % (rc("status"), rc("scope_id"), P1, wc, wa, we, wk))
    bad = ('SUMPRODUCT((%s="open")*(%s=$B$1)*((%s="")+(%s="")+(%s="")))'
           % (r("status"), r("scope_id"), r("criticality_tier"),
              r("first_approval_date"), r("exception_scope")))
    return dict(columns=cols, params=["scope"], helpers=[helper],
                outputs={"value": "=IF(%s>0,NA(),SUM(%s))" % (bad, H)},
                note="Helper column weighted_exposure applies the card's banded weights "
                     "(tier x age x extent x compensating control) per open exception; value = the sum. "
                     "Zero open exceptions is a true 0; missing tier/date/extent on an open row -> n/a.")


# curated flagships with card-specific semantics + the three Stage-3 specials
_SPECIAL = {
    "STD-016": lambda cols, card, k: _ratio_concrete_std016(cols),
    "SOC-002": lambda cols, card, k: _duration(cols, "occurred_at", "detected_at", True, True),
    "STD-001": lambda cols, card, k: _std001(cols),
    "RES-007": lambda cols, card, k: _res007(cols),
    "LOG-015": lambda cols, card, k: _log015(cols),
}


def build_spec(card, recipe, comp_k=None):
    """Deterministic Excel spec from the card contract + generated recipe. None -> no xlsx dialect."""
    cid = card["id"]; cols = card["minimum_data_fields"]; mech = recipe.get("mechanic")
    if cid in _SPECIAL:
        return _SPECIAL[cid](cols, card, comp_k)
    status = recipe.get("status")
    if status == "pending" or mech == "ranking":
        return None
    if mech == "ratio":
        m = RATIO_PAT.match((card.get("formula") or "").strip())
        nd = (card.get("numerator_denominator") or "").strip()
        hooks = ({"num": m.group(1).replace("_", " "), "den": m.group(2).replace("_", " ")} if m else
                 {"num": (nd.split("/")[0].strip() or "numerator per the card definition"),
                  "den": (nd.split("/")[1].strip() if "/" in nd else "denominator per the card definition")})
        return _ratio_skeleton(cid, cols, hooks)
    if mech == "duration":
        m = DUR_PAT.search(card.get("formula") or "")
        concrete = bool(m and m.group(1) in cols and m.group(2) in cols)
        end, start = (m.group(1), m.group(2)) if concrete else ("end_at", "start_at")
        return _duration(cols, start, end, False, concrete)
    if mech == "count":
        return _count_skeleton(cid, cols, ((card.get("numerator_denominator") or card.get("formula") or "")[:110]).strip())
    if mech == "delta":
        return _delta_skeleton(cid, cols, ((card.get("numerator_denominator") or card.get("formula") or "")[:110]).strip())
    if mech == "component_tree" and comp_k:
        return _composite(cols, comp_k)
    return None


# ============ display rendering ============
def render(cid, spec):
    out = ["' %s \u00b7 Excel recipe (range model) \u00b7 non-normative" % cid]
    pm = {"scope": "B1 = scope", "ps": "B2 = period start", "pe": "B3 = period end"}
    out.append("' On a sheet named 'result' set the parameters: "
               + "; ".join(pm[p] for p in spec["params"]))
    out.append("' Put raw rows on a sheet named 'data' (header in row 1), columns in this order:")
    out.append("'   " + ", ".join(spec["columns"]))
    for h, (hname, hform) in enumerate(spec["helpers"]):
        hcol = LETTER(len(spec["columns"]) + 1 + h)
        out.append("' Helper column %s (paste in %s2 and fill down beside your data):" % (hname, hcol))
        out.append("    " + hform.format(r=2))
    out.append("' Result cell(s):")
    for key, form in spec["outputs"].items():
        out.append("    %s%s" % (key + ":  " if key != "value" else "", form))
    out.append("' German Excel (de-DE) - same formula, localised names, ';' separator, ',' decimal;")
    out.append("' copy the line after the apostrophe (a saved .xlsx localises automatically on open):")
    for key, form in spec["outputs"].items():
        out.append("'   %s%s" % (key + ":  " if key != "value" else "", to_de_de(form)))
    return "\n".join(out)


# ============ executable workbook build + dual-engine eval (shared by gate & runner) ============
def _parse_dt(v):
    if isinstance(v, str) and re.match(r"\d{4}-\d{2}-\d{2}T", v):
        return _dt.datetime.strptime(v, "%Y-%m-%dT%H:%M:%SZ")
    return v

def build_workbook(spec, fixture, path):
    import openpyxl
    wb = openpyxl.Workbook(); res = wb.active; res.title = "result"; data = wb.create_sheet("data")
    cols = spec["columns"]; rows = (fixture or {}).get("rows", []); N = len(rows)
    for j, f in enumerate(cols, 1): data.cell(1, j, f)
    for i, row in enumerate(rows, start=2):
        for j, f in enumerate(cols, 1):
            v = row.get(f)
            if v is not None: data.cell(i, j, _parse_dt(v))
    for h, (hname, hform) in enumerate(spec["helpers"]):
        hc = len(cols) + 1 + h; data.cell(1, hc, hname)
        for i in range(2, max(N, 1) + 2):
            data.cell(i, hc, hform.format(r=i))
    params = (fixture or {}).get("params", {})
    res["A1"] = "scope"; res["B1"] = params.get("scope_id", "prod")
    res["A2"] = "period_start"; res["B2"] = _parse_dt(params.get("period_start", "2026-06-01T00:00:00Z"))
    res["A3"] = "period_end"; res["B3"] = _parse_dt(params.get("period_end", "2026-07-01T00:00:00Z"))
    out_rows = {}
    for k, (key, form) in enumerate(spec["outputs"].items()):
        rr = 6 + k; res.cell(rr, 1, key); res.cell(rr, 2, form); out_rows[key] = rr
    wb.move_sheet("result", -(len(wb.sheetnames) - 1))
    wb.save(path)
    return out_rows

def _cellnum(x):
    if x is None: return None
    s = str(x).strip()
    if s in ("#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#REF!", ""): return s
    try: return float(s)
    except Exception: return s

def eval_formulas(path, out_rows):
    import warnings; warnings.filterwarnings("ignore")
    import formulas
    xl = formulas.ExcelModel().loads(path).finish(); sol = xl.calculate()
    up = {k.upper(): v for k, v in sol.items()}
    res = {}
    for key, rr in out_rows.items():
        tgt = ("RESULT'!B%d" % rr).upper(); val = None
        for k, v in up.items():
            if k.endswith(tgt):
                try: val = v.value[0, 0]
                except Exception: val = getattr(v, "value", v)
                break
        res[key] = _cellnum(val)
    return res

def eval_libreoffice(path, out_rows, soffice="soffice", timeout=120, user_installation=None):
    outdir = os.path.dirname(path) or "."
    base = os.path.splitext(os.path.basename(path))[0]
    csv = os.path.join(outdir, base + ".csv")
    if os.path.exists(csv): os.remove(csv)
    cmd = [soffice, "--headless"]
    if user_installation:
        cmd.append("-env:UserInstallation=%s" % user_installation)
    cmd += ["--convert-to", "csv:Text - txt - csv (StarCalc)", "--outdir", outdir, path]
    subprocess.run(cmd, capture_output=True, timeout=timeout)
    label_val = {}
    if os.path.exists(csv):
        for ln in open(csv, encoding="utf-8").read().splitlines():
            parts = ln.split(",")
            if len(parts) >= 2 and parts[0]:
                label_val[parts[0]] = parts[1]
    return {key: _cellnum(label_val.get(key)) for key in out_rows}
